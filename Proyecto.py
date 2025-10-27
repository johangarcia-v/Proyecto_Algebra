import pygame
import random
import sys
import time
import os
import math
from pygame import mixer
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Para exportar a Excel, instala openpyxl con: pip install openpyxl")

# Inicializar Pygame y el mixer para el sonido
pygame.init()
mixer.init()

# Configuración del juego
NUM_COLORS = 6  # Número de colores diferentes para los caramelos
COLOR_MAP = [
    (255, 0, 0),     # Rojo
    (0, 255, 0),     # Verde
    (0, 0, 255),     # Azul
    (255, 255, 0),   # Amarillo
    (255, 0, 255),   # Magenta
    (0, 255, 255),   # Cian
]

# Configuración de la ventana y juego
class Config:
    # Dimensiones de la ventana
    ANCHO = 1200
    ALTO = 800
    
    # Configuración del tablero
    TAMANO_CELDA = 64
    FILAS = 8
    COLUMNAS = 8
    
    # Configuración de paneles
    RIGHT_PANEL_WIDTH = 360
    LEFT_WIDTH = ANCHO - RIGHT_PANEL_WIDTH
    
    # Márgenes calculados
    MARGEN_SUPERIOR = (ALTO - (FILAS * TAMANO_CELDA)) // 2
    MARGEN_IZQUIERDO = (LEFT_WIDTH - (COLUMNAS * TAMANO_CELDA)) // 2
    
    # Colores
    BLANCO = (255, 255, 255)
    NEGRO = (0, 0, 0)
    AZUL = (10, 90, 180)  # Fondo derecho (Uniminuto azul aproximado)
    GRIS = (230, 230, 230)

# Usar las constantes desde la clase Config
ANCHO = Config.ANCHO
ALTO = Config.ALTO
TAMANO_CELDA = Config.TAMANO_CELDA
FILAS = Config.FILAS
COLUMNAS = Config.COLUMNAS
RIGHT_PANEL_WIDTH = Config.RIGHT_PANEL_WIDTH
LEFT_WIDTH = Config.LEFT_WIDTH
MARGEN_SUPERIOR = Config.MARGEN_SUPERIOR
MARGEN_IZQUIERDO = Config.MARGEN_IZQUIERDO
BLANCO = Config.BLANCO
NEGRO = Config.NEGRO
AZUL = Config.AZUL
GRIS = Config.GRIS

# Colores
BLANCO = (255, 255, 255)
NEGRO = (0, 0, 0)
AZUL = (10, 90, 180)  # Fondo derecho (Uniminuto azul aproximado)
GRIS = (230, 230, 230)

# Configurar la ventana
ventana = pygame.display.set_mode((ANCHO, ALTO))
pygame.display.set_caption('Candy Matrix - Proyecto Algebra Lineal')

# Gestor de sonidos del juego
class SoundManager:
    def __init__(self):
        self.sounds = {}
        self._load_sounds()
        self._setup_music()
    
    def _load_sounds(self):
        sound_files = {
            'explosion': os.path.join('sounds', 'bubble-pop-06-351337.mp3')
        }
        
        for name, path in sound_files.items():
            if os.path.exists(path):
                try:
                    self.sounds[name] = mixer.Sound(path)
                except Exception as e:
                    print(f'No se pudo cargar el sonido {name}:', e)
    
    def _setup_music(self):
        music_path = os.path.join('sounds', 'background_music.mp3')
        if os.path.exists(music_path):
            try:
                mixer.music.load(music_path)
                mixer.music.set_volume(0.5)
                mixer.music.play(-1)
            except Exception as e:
                print('No se pudo reproducir la música:', e)
    
    def play_sound(self, name):
        if name in self.sounds:
            self.sounds[name].play()

# Inicializar el gestor de sonidos
sound_manager = SoundManager()
def export_to_excel():
    if not EXCEL_AVAILABLE:
        print("Necesitas instalar openpyxl para exportar a Excel")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumen del Juego"

    # Estilos
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = "CANDY MATRIX - RESUMEN DEL JUEGO"
    ws['A1'].font = title_font
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal='center')

    # Datos principales con explicaciones en la columna A/B (campo/valor) y C (explicación)
    info_rows = [
        ("Fecha", datetime.now().strftime("%Y-%m-%d %H:%M"), "Momento en que se generó el reporte"),
        ("Nivel alcanzado", globals().get('level', 'N/A'), "Nivel alcanzado por el jugador al generar el reporte"),
        ("Puntuación final", globals().get('score', 0), "Puntos totales obtenidos en la sesión"),
        ("Movimientos totales", globals().get('moves_count', 0), "Número de intercambios/movimientos realizados")
    ]

    start_row = 3
    for i, (campo, valor, explicacion) in enumerate(info_rows, start_row):
        ws.cell(row=i, column=1, value=campo)
        ws.cell(row=i, column=2, value=valor)
        ws.cell(row=i, column=3, value=explicacion)
        # aplicar estilo a la fila de campo
        ws.cell(row=i, column=1).font = header_font
        ws.cell(row=i, column=1).fill = header_fill
        ws.cell(row=i, column=1).alignment = openpyxl.styles.Alignment(horizontal='left')

    # Estadísticas de matriz actual
    stats = matrix_stats()
    stats_row = start_row + len(info_rows) + 2
    ws.cell(row=stats_row, column=1, value="ESTADÍSTICAS DE MATRIZ FINAL").font = title_font
    stat_items = [
        ("Suma total", stats['sum'], "Suma de todos los valores numéricos en la matriz"),
        ("Promedio", round(stats['mean'], 2), "Promedio de los valores numéricos"),
        ("Mínimo", stats['min'], "Valor mínimo presente en la matriz"),
        ("Máximo", stats['max'], "Valor máximo presente en la matriz")
    ]
    for idx, (campo, valor, explicacion) in enumerate(stat_items, 1):
        r = stats_row + idx
        ws.cell(row=r, column=1, value=campo)
        ws.cell(row=r, column=2, value=valor)
        ws.cell(row=r, column=3, value=explicacion)

    # Historial de matrices y eventos en hoja separada
    ws2 = wb.create_sheet(title="Historial de Matrices")
    ws2.cell(row=1, column=1, value="Movimiento").font = header_font
    ws2.cell(row=1, column=2, value="Evento").font = header_font
    ws2.cell(row=1, column=3, value="Matriz (filas a continuación)").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3).fill = header_fill

    row = 2
    # Asegurar que las estructuras existen
    hist = globals().get('matrix_history', [])
    events = globals().get('game_events', [])
    for idx, matrix in enumerate(hist, 1):
        evento = events[idx-1] if idx-1 < len(events) else ''
        ws2.cell(row=row, column=1, value=f"Movimiento {idx}")
        ws2.cell(row=row, column=2, value=evento)
        # Escribir la matriz debajo como bloque
        row += 1
        for i, matrix_row in enumerate(matrix):
            for j, val in enumerate(matrix_row):
                ws2.cell(row=row+i, column=j+1, value=("#" if val is None else val))
        row += len(matrix) + 1

    # Hoja de explicación con descripciones de columnas y uso del archivo
    ws3 = wb.create_sheet(title="Explicación")
    ws3.cell(row=1, column=1, value="Campo").font = header_font
    ws3.cell(row=1, column=2, value="Descripción").font = header_font
    ws3.cell(row=1, column=1).fill = header_fill
    ws3.cell(row=1, column=2).fill = header_fill
    explanations = [
        ("Fecha", "Fecha y hora en que se generó el reporte"),
        ("Nivel alcanzado", "Nivel del jugador en el momento del reporte"),
        ("Puntuación final", "Puntos acumulados en la sesión"),
        ("Movimientos totales", "Número total de movimientos/intercambios realizados"),
        ("ESTADÍSTICAS DE MATRIZ FINAL", "Resumen estadístico de la matriz actual: suma, promedio, min y max"),
        ("Historial de Matrices", "Registros de movimientos con la matriz tras cada movimiento para análisis")
    ]
    for i, (campo, desc) in enumerate(explanations, 2):
        ws3.cell(row=i, column=1, value=campo)
        ws3.cell(row=i, column=2, value=desc)

    # Ajustar anchos de columnas en la hoja principal y hojas auxiliares
    from openpyxl.utils import get_column_letter
    for sheet in [ws, ws2, ws3]:
        for col in range(1, sheet.max_column + 1):
            col_letter = get_column_letter(col)
            max_length = 0
            for cell in sheet[col_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            sheet.column_dimensions[col_letter].width = min(50, max_length + 4)

    # Guardar archivo en la carpeta Descargas del usuario (Windows/Linux/Mac)
    downloads = os.path.join(os.path.expanduser('~'), 'Downloads')
    if not os.path.isdir(downloads):
        # Si no existe Downloads, usar el directorio actual
        downloads = os.getcwd()
    filename = f"candy_matrix_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    fullpath = os.path.join(downloads, filename)
    wb.save(fullpath)
    return fullpath
MAX_LEVEL = 5
LEVEL_TIME = 180  # 3 minutos
level_start_time = time.time()
goal_base = 500

# Tiempos globales
game_start_time = time.time()

# Animaciones sencillas: registrar últimas explosiones [(pos_list, start_time)]
last_explosions = []
# Hover
hover_cell = None

# Configuración de habilidades
class Habilidad:
    BOMBA = 'bomb'
    ARCOIRIS = 'rainbow'
    ESTRELLA = 'star'
    
    @staticmethod
    def get_all_types():
        return [Habilidad.BOMBA, Habilidad.ARCOIRIS, Habilidad.ESTRELLA]

# Habilidades en el mapa: dict (fila,col) -> tipo ('bomb','rainbow','star')
skills_on_map = {}

# Preguntas True/False sobre álgebra lineal
questions = [
    ("La suma de matrices A+B es conmutativa", True),
    ("El producto de matrices AB = BA siempre", False),
    ("Una matriz singular es no invertible", True),
    ("El rango de una matriz es <= min(m,n)", True),
    ("Toda matriz es diagonalizable", False),
    ("tr(AB) = tr(BA) para toda matriz", True),
    ("Los autovalores son únicos", True),
    ("La traza es la suma de autovalores", True),
    ("Una matriz simétrica es diagonalizable", True),
    ("det(AB) = det(A)det(B)", True),
    ("La inversa de una matriz ortogonal es su transpuesta", True),
    ("Una matriz triangular superior no tiene autovalores", False),
    ("El espacio nulo es un subespacio vectorial", True),
    ("Los autovectores de valores distintos son LI", True),
    ("Una matriz idempotente cumple A^2 = A", True)
]
# Preguntas y cooldowns
quiz_cooldown_until = 0
# cooldown después de respuesta correcta (timestamp hasta el que no se puede usar)
skill_success_cooldown_until = 0
# límite de usos de habilidades por nivel
MAX_SKILL_USES_PER_LEVEL = 4
skill_uses_this_level = 0
asked_questions = set()

# --- Nuevas funciones para preguntas MCQ / basadas en la matriz ---
def make_mcq(question_text, correct_answer, distractors=None):
    """Crear estructura MCQ con opciones mezcladas y el índice correcto."""
    if distractors is None:
        distractors = []
    options = [str(correct_answer)] + [str(d) for d in distractors]
    random.shuffle(options)
    correct_idx = options.index(str(correct_answer))
    # dividir el texto en líneas para render
    words = question_text.split()
    lines = []
    cur = words[0] if words else ''
    for w in words[1:]:
        if len(cur + ' ' + w) <= 40:
            cur += ' ' + w
        else:
            lines.append(cur)
            cur = w
    if cur:
        lines.append(cur)
    return {'lines': lines, 'options': options, 'correct_idx': correct_idx}

def generate_matrix_mcq():
    """Genera una pregunta basada en la matriz actual: suma fila, suma columna o valor de celda."""
    typ = random.choice(['row_sum', 'col_sum', 'cell_val'])
    if typ == 'row_sum':
        r = random.randrange(FILAS)
        correct = sum(v for v in tablero[r] if isinstance(v, int) and v >= 0)
        # crear distractores cercanos
        distractors = [correct + random.randint(-3, -1), correct + random.randint(1, 4)]
        q = f'¿Cuál es la suma de los valores en la fila {r} de la matriz?'
        return make_mcq(q, correct, distractors)
    elif typ == 'col_sum':
        c = random.randrange(COLUMNAS)
        colvals = [tablero[i][c] for i in range(FILAS) if isinstance(tablero[i][c], int) and tablero[i][c] >= 0]
        correct = sum(colvals)
        distractors = [correct + random.randint(-4, -1), correct + random.randint(1, 5)]
        q = f'¿Cuál es la suma de los valores en la columna {c} de la matriz?'
        return make_mcq(q, correct, distractors)
    else:
        i = random.randrange(FILAS)
        j = random.randrange(COLUMNAS)
        val = tablero[i][j]
        correct = val if isinstance(val, int) and val >= 0 else 0
        distractors = [correct + 1, correct - 1]
        q = f'¿Cuál es el valor en la celda ({i},{j}) de la matriz?'
        return make_mcq(q, correct, distractors)

def draw_row_sums_bar(surface, area_rect):
    """Dibuja un gráfico de barras simple con las sumas de cada fila del tablero."""
    row_sums = [sum(v for v in row if isinstance(v, int) and v >= 0) for row in tablero]
    if not row_sums:
        return
    maxv = max(row_sums)
    padding = 8
    gw = area_rect.width - 2*padding
    gh = area_rect.height - 2*padding
    bar_w = gw / len(row_sums) - 6
    for i, val in enumerate(row_sums):
        x = area_rect.x + padding + i * (bar_w + 6)
        h = int((val / maxv) * (gh - 20)) if maxv > 0 else 0
        y = area_rect.y + area_rect.height - padding - h
        pygame.draw.rect(surface, (100,180,220), (x, y, int(bar_w), h))
        # etiqueta
        lbl = font.render(str(val), True, (20,20,20))
        surface.blit(lbl, (x, y - 18))
    # título pequeño
    surface.blit(font.render('Suma por fila', True, (10,10,10)), (area_rect.x + padding, area_rect.y + 4))

def draw_transform_demo(surface, area_rect):
    """Dibuja una demo simple de transformación 2D aplicando una matriz a un cuadro unitario."""
    # matriz ejemplo
    M = [[1.2, 0.6], [-0.5, 1.0]]
    # centro del área
    cx = area_rect.x + area_rect.width // 2
    cy = area_rect.y + area_rect.height // 2
    scale = min(area_rect.width, area_rect.height) * 0.18
    # puntos del cuadrado unidad
    pts = [(0,0), (1,0), (1,1), (0,1), (0,0)]
    # dibujar eje original
    pygame.draw.line(surface, (180,180,180), (cx - scale*2, cy), (cx + scale*2, cy), 1)
    pygame.draw.line(surface, (180,180,180), (cx, cy - scale*2), (cx, cy + scale*2), 1)
    # transformar y dibujar
    transformed = []
    for (x,y) in pts:
        tx = M[0][0]*x + M[0][1]*y
        ty = M[1][0]*x + M[1][1]*y
        sx = cx + tx * scale
        sy = cy - ty * scale
        transformed.append((int(sx), int(sy)))
    pygame.draw.lines(surface, (220,100,80), False, transformed, 3)
    # dibujo de origen en semitransparente
    orig = []
    for (x,y) in pts:
        ox = cx + x * scale
        oy = cy - y * scale
        orig.append((int(ox), int(oy)))
    pygame.draw.lines(surface, (120,120,120), False, orig, 1)
    surface.blit(font.render('Transformación 2D (ejemplo)', True, (10,10,10)), (area_rect.x + 6, area_rect.y + 4))

# Pequeña colección de páginas para la mini-lección
LEARN_PAGES = [
    ("Matrices y operaciones", [
        "Una matriz es una tabla de números organizada en filas y columnas.",
        "Operaciones básicas: suma, resta, multiplicación y trasposición.",
        "Observa cómo las operaciones cambian la estructura: prueba con matrices pequeñas."
    ]),
    ("Rango y soluciones", [
        "El rango indica cuántas columnas/filas son linealmente independientes.",
        "Si det(A) != 0 para matriz cuadrada, A es invertible y la solución es única.",
        "Autovalores/autovectores: vectores que sólo cambian de escala al aplicar A."
    ]),
    ("Ejemplos prácticos", [
        "Calcula determinantes de 2x2 manualmente: det([[a,b],[c,d]]) = ad - bc.",
        "Usa el juego para relacionar números de la matriz con preguntas concretas.",
        "Intenta explicar por qué una matriz singular no tiene inversa."
    ]),
    ("Consejos de estudio", [
        "Practica con ejercicios cortos y verifica con el juego para feedback inmediato.",
        "Haz tarjetas (flashcards) de definiciones clave: traza, rango, determinante.",
        "Divide el estudio: 20 min teoría + 20 min ejercicios prácticos."
    ])
]

def show_learn_screen():
    """Pantalla de mini-lección modal con navegación entre páginas."""
    global level_start_time  # Para poder ajustar el tiempo del nivel
    
    # Guardar tiempo actual para ajustarlo después
    tiempo_pausa = time.time()
    
    page = 0
    clock = pygame.time.Clock()
    # Loop modal: procesa eventos y dibuja una ventana central con contenido y visualizaciones
    while True:
        # Preparar rects de botones (mantenerlos dentro de la ventana)
        w = 900
        h = 520
        rx = (ANCHO - w)//2
        ry = (ALTO - h)//2
        rect = pygame.Rect(rx, ry, w, h)
        btn_prev = pygame.Rect(rect.x + 20, rect.y + h - 64, 120, 40)
        btn_next = pygame.Rect(rect.x + rect.width - 140, rect.y + h - 64, 120, 40)
        btn_close = pygame.Rect(rect.x + rect.width//2 - 60, rect.y + h - 64, 120, 40)

        for ev in pygame.event.get():
            if ev.type == pygame.QUIT:
                pygame.quit(); sys.exit()
            if ev.type == pygame.KEYDOWN:
                if ev.key == pygame.K_ESCAPE:
                    # Ajustar el tiempo del nivel para compensar la pausa
                    tiempo_actual = time.time()
                    tiempo_pausado = tiempo_actual - tiempo_pausa
                    level_start_time += tiempo_pausado
                    return
                if ev.key == pygame.K_RIGHT:
                    page = min(page + 1, len(LEARN_PAGES) - 1)
                if ev.key == pygame.K_LEFT:
                    page = max(page - 1, 0)
            if ev.type == pygame.MOUSEBUTTONDOWN:
                mx, my = pygame.mouse.get_pos()
                if btn_next.collidepoint(mx, my):
                    page = min(page + 1, len(LEARN_PAGES) - 1)
                if btn_prev.collidepoint(mx, my):
                    page = max(page - 1, 0)
                if btn_close.collidepoint(mx, my):
                    # Ajustar el tiempo del nivel para compensar la pausa
                    tiempo_actual = time.time()
                    tiempo_pausado = tiempo_actual - tiempo_pausa
                    level_start_time += tiempo_pausado
                    return

        # Dibujar fondo semi-transparente y ventana central
        s = pygame.Surface((ANCHO, ALTO), pygame.SRCALPHA)
        s.fill((10, 10, 10, 200))
        ventana.blit(s, (0, 0))
        pygame.draw.rect(ventana, (245,245,255), rect, border_radius=12)
        pygame.draw.rect(ventana, (200,200,200), rect, 2, border_radius=12)

        title, lines = LEARN_PAGES[page]
        ventana.blit(large_font.render(title, True, (20,20,20)), (rect.x + 24, rect.y + 18))
        for i, l in enumerate(lines):
            ventana.blit(font.render('- ' + l, True, (40,40,40)), (rect.x + 24, rect.y + 70 + i*30))

        # Renderizar el contenido textual
        y_offset = rect.y + 70
        for l in lines:
            ventana.blit(font.render('- ' + l, True, (40,40,40)), (rect.x + 24, y_offset))
            y_offset += 30

        # Añadir espacio adicional antes de las visualizaciones
        y_offset += 40
        
        # Visualizaciones centradas debajo del texto
        viz_area = pygame.Rect(rect.x + (rect.width - 400)//2, y_offset + 40, 400, 160)
        if title == 'Matrices y operaciones':
            # Texto adicional antes del gráfico
            extra_text = ["La suma por fila nos ayuda a visualizar la distribución de valores",
                         "y detectar patrones en la estructura de la matriz."]
            for l in extra_text:
                ventana.blit(font.render(l, True, (40,40,40)), (rect.x + 24, y_offset))
                y_offset += 25
            y_offset += 30  # Más espacio antes del gráfico
            draw_row_sums_bar(ventana, viz_area)
        elif title == 'Rango y soluciones':
            draw_transform_demo(ventana, viz_area)
        elif title == 'Ejemplos prácticos':
            m2x2 = [[1, 2], [3, 4]]
            det = m2x2[0][0]*m2x2[1][1] - m2x2[0][1]*m2x2[1][0]
            ventana.blit(font.render('Ejemplo 2x2:', True, (20,20,20)), (viz_area.x + 6, viz_area.y + 6))
            ventana.blit(font.render(f'[{m2x2[0][0]} {m2x2[0][1]}]', True, (20,20,20)), (viz_area.x + 6, viz_area.y + 36))
            ventana.blit(font.render(f'[{m2x2[1][0]} {m2x2[1][1]}]', True, (20,20,20)), (viz_area.x + 6, viz_area.y + 60))
            ventana.blit(font.render(f'det = {det}', True, (120,10,10)), (viz_area.x + 6, viz_area.y + 96))

        # Botones más pequeños y mejor espaciados
        btn_width = 100
        btn_height = 35
        spacing = 20
        
        # Recalcular posiciones de botones
        btn_prev = pygame.Rect(rect.x + spacing, rect.y + h - 50, btn_width, btn_height)
        btn_next = pygame.Rect(rect.x + rect.width - btn_width - spacing, rect.y + h - 50, btn_width, btn_height)
        btn_close = pygame.Rect(rect.x + (rect.width - btn_width)//2, rect.y + h - 50, btn_width, btn_height)
        
        # Dibujar botones con nuevo estilo
        pygame.draw.rect(ventana, (200,200,200), btn_prev, border_radius=6)
        pygame.draw.rect(ventana, (200,200,200), btn_next, border_radius=6)
        pygame.draw.rect(ventana, (180, 120, 120), btn_close, border_radius=6)
        
        # Centrar texto en los botones
        for btn, texto in [(btn_prev, 'Anterior'), (btn_next, 'Siguiente'), (btn_close, 'Cerrar')]:
            txt_surf = font.render(texto, True, (0,0,0))
            txt_rect = txt_surf.get_rect(center=btn.center)
            ventana.blit(txt_surf, txt_rect)

        pygame.display.flip()
        clock.tick(30)

# Popups flotantes no bloqueantes (nivel up)
floating_popups = []

font = pygame.font.SysFont('arial', 20)
large_font = pygame.font.SysFont('arial', 30)


def ellipsize(text, font_obj, max_width):
    """Recorta el texto con '...' si excede el ancho máximo en píxeles."""
    if font_obj.size(text)[0] <= max_width:
        return text
    # dejar espacio para '...'
    ell = '...'
    max_w = max_width - font_obj.size(ell)[0]
    if max_w <= 0:
        return ell
    # recortar iterativamente (simple pero eficaz para strings cortos)
    for length in range(len(text), 0, -1):
        candidate = text[:length]
        if font_obj.size(candidate)[0] <= max_w:
            return candidate + ell
    return ell

def dibujar_interface():
    # Panel izquierdo: fondo tipo canasta de dulces (tono más claro)
    ventana.fill((245, 230, 235))
    # Dibujar panel principal (canasta)
    panel_rect = pygame.Rect(10, 10, LEFT_WIDTH-20, ALTO-20)
    # Fondo de madera claro (menos saturado)
    pygame.draw.rect(ventana, (195, 155, 110), panel_rect, border_radius=20)
    # Tejido de la canasta (patrón entrecruzado)
    stripe_color1 = (160, 82, 45)
    stripe_color2 = (205, 133, 63)
    spacing = 15
    # Líneas verticales
    for x in range(panel_rect.x, panel_rect.x + panel_rect.width, spacing):
        pygame.draw.line(ventana, stripe_color1, (x, panel_rect.y), (x, panel_rect.y + panel_rect.height), 2)
    # Líneas horizontales entrelazadas
    for y in range(panel_rect.y, panel_rect.y + panel_rect.height, spacing*2):
        for x in range(panel_rect.x, panel_rect.x + panel_rect.width - spacing, spacing*2):
            pygame.draw.rect(ventana, stripe_color2, (x, y, spacing, spacing))
    # Borde superior reforzado
    pygame.draw.rect(ventana, stripe_color2, (panel_rect.x, panel_rect.y, panel_rect.width, spacing*2), border_radius=20)

    # Panel derecho (azul)
    pygame.draw.rect(ventana, AZUL, (LEFT_WIDTH, 0, RIGHT_PANEL_WIDTH, ALTO))

    # Dibujar caramelos (círculos) en la cuadrícula centrada
    for i in range(FILAS):
        for j in range(COLUMNAS):
            # Coordenadas centrales de la celda
            x = MARGEN_IZQUIERDO + j * TAMANO_CELDA + TAMANO_CELDA // 2
            y = MARGEN_SUPERIOR + i * TAMANO_CELDA + TAMANO_CELDA // 2
            radius = TAMANO_CELDA // 2 - 8

            # Dibujar fondo de celda (sutil)
            cell_bg = pygame.Surface((TAMANO_CELDA-8, TAMANO_CELDA-8), pygame.SRCALPHA)
            pygame.draw.rect(cell_bg, (0,0,0,30), (0,0,TAMANO_CELDA-8,TAMANO_CELDA-8), border_radius=8)
            ventana.blit(cell_bg, (x - (TAMANO_CELDA-8)//2, y - (TAMANO_CELDA-8)//2))

            # Dibujar caramelo si existe
            if tablero[i][j] is not None:
                color = COLOR_MAP[tablero[i][j]]
                # Sombras y borde para aspecto dulce
                pygame.draw.circle(ventana, (80, 50, 60), (x+2, y+4), radius+2)
                pygame.draw.circle(ventana, (230, 230, 230), (x, y), radius+1)
                pygame.draw.circle(ventana, color, (x, y), radius)
                # Brillo superior izquierdo
                shine = pygame.Surface((radius*2, radius*2), pygame.SRCALPHA)
                pygame.draw.ellipse(shine, (255,255,255,90), (int(radius*0.1), int(radius*0.05), int(radius*0.9), int(radius*0.6)))
                ventana.blit(shine, (x - radius, y - radius))

            # Dibujar habilidades como caramelos especiales (si hay)
            if (i, j) in skills_on_map:
                val = skills_on_map[(i, j)]
                if isinstance(val, tuple):
                    typ, orig_color = val
                else:
                    typ = val
                    orig_color = None
                if typ == 'bomb':
                    # Caramelo bomba (negro con detalles rojos)
                    pygame.draw.circle(ventana, (40, 40, 40), (x, y), radius)
                    # Detalles de la mecha
                    pygame.draw.line(ventana, (255, 50, 50), (x, y-radius+2), (x+4, y-radius-6), 3)
                    pygame.draw.circle(ventana, (255, 200, 50), (x+4, y-radius-6), 3)
                    # Brillo
                    shine = pygame.Surface((radius*2, radius*2), pygame.SRCALPHA)
                    pygame.draw.ellipse(shine, (255,255,255,40), (radius*0.5, radius*0.5, radius, radius))
                    ventana.blit(shine, (x - radius, y - radius))
                elif typ == 'rainbow':
                    # Caramelo arcoíris (efecto giratorio)
                    t = time.time() * 2
                    for angle in range(0, 360, 30):
                        rad = math.radians(angle + t * 30)
                        color = pygame.Color(0)
                        color.hsva = (angle % 360, 100, 100, 100)
                        start = (x + math.cos(rad)*radius*0.5, y + math.sin(rad)*radius*0.5)
                        end = (x + math.cos(rad)*radius*0.9, y + math.sin(rad)*radius*0.9)
                        pygame.draw.line(ventana, color, start, end, 3)
                    # Centro blanco
                    pygame.draw.circle(ventana, (255,255,255), (x, y), int(radius*0.4))
                elif typ == 'star':
                    # Caramelo estrella (dorado brillante)
                    pygame.draw.circle(ventana, (255, 215, 0), (x, y), radius)
                    # Destellos de estrella
                    t = time.time() * 3
                    for angle in range(0, 360, 45):
                        rad = math.radians(angle + t * 30)
                        length = radius * (0.5 + math.sin(t*2 + angle*0.1) * 0.2)
                        end_x = x + math.cos(rad) * length
                        end_y = y + math.sin(rad) * length
                        pygame.draw.line(ventana, (255, 255, 200), (x, y), (end_x, end_y), 2)
                    # Brillo central
                    shine = pygame.Surface((radius*2, radius*2), pygame.SRCALPHA)
                    pygame.draw.ellipse(shine, (255,255,255,128), (radius*0.6, radius*0.6, radius*0.8, radius*0.8))
                    ventana.blit(shine, (x - radius, y - radius))

    # Resaltar seleccionado
    if seleccionado:
        i, j = seleccionado
        x = MARGEN_IZQUIERDO + j * TAMANO_CELDA
        y = MARGEN_SUPERIOR + i * TAMANO_CELDA
        pygame.draw.rect(ventana, BLANCO, (x-3, y-3, TAMANO_CELDA+6, TAMANO_CELDA+6), 3, border_radius=10)

    # Resaltar hover
    if hover_cell:
        hi, hj = hover_cell
        if 0 <= hi < FILAS and 0 <= hj < COLUMNAS:
            hx = MARGEN_IZQUIERDO + hj * TAMANO_CELDA
            hy = MARGEN_SUPERIOR + hi * TAMANO_CELDA
            s = pygame.Surface((TAMANO_CELDA-4, TAMANO_CELDA-4), pygame.SRCALPHA)
            s.fill((255,255,255,30))
            ventana.blit(s, (hx, hy))

    # Dibujar animaciones de explosión recientes
    now = time.time()
    new_expl = []
    for cells, t0 in last_explosions:
        age = now - t0
        if age > 0.6:
            continue
        new_expl.append((cells, t0))
        # animación simple: círculo con alpha que crece
        for (i, j) in cells:
            cx = MARGEN_IZQUIERDO + j * TAMANO_CELDA + TAMANO_CELDA//2 - 2
            cy = MARGEN_SUPERIOR + i * TAMANO_CELDA + TAMANO_CELDA//2 - 2
            radius = int(6 + age * 30)
            alpha = int(200 * (1 - age/0.6))
            surf = pygame.Surface((radius*2, radius*2), pygame.SRCALPHA)
            pygame.draw.circle(surf, (255,255,255,alpha), (radius, radius), radius)
            ventana.blit(surf, (cx - radius + 2, cy - radius + 2))
    # mantener solo las no expiradas
    last_explosions[:] = new_expl

    # Panel derecho: mostrar matriz numérica y estadísitcas
    draw_right_panel()

    # Dibujar popups flotantes (no bloqueantes)
    now = time.time()
    to_remove = []
    for idx, p in enumerate(floating_popups):
        age = now - p['start']
        if age > p['duration']:
            to_remove.append(idx)
            continue
        # calcular posición en y (subiendo)
        y = p['y'] + p['vy'] * age
        # calcular tamaño dinámico según el texto para que el fondo lo cubra completamente
        padding_x = 12
        padding_y = 8
        max_popup_w = min(ANCHO - 40, 520)

        # soportar icono opcional
        icon = p.get('icon', None)
        icon_space = 0
        icon_size = 0
        if icon:
            icon_size = 18
            icon_space = icon_size + 8

        # Limitar y recortar líneas muy largas
        max_text_w_allowed = max_popup_w - padding_x*2 - icon_space
        rendered_lines = []
        for line in p['lines']:
            short = ellipsize(line, font, max_text_w_allowed)
            rendered_lines.append(short)

        line_surfs = [font.render(line, True, (255,255,255)) for line in rendered_lines]
        max_text_w = max((surf.get_width() for surf in line_surfs), default=0)
        total_text_h = sum(surf.get_height() for surf in line_surfs) + (len(line_surfs)-1) * 6

        w = max(200, max_text_w + padding_x * 2 + icon_space)
        w = min(w, max_popup_w)
        h = total_text_h + padding_y * 2

        rect_x = p.get('x', ANCHO//2 - w//2)
        # si la caja sale de la pantalla a la derecha, ajustarla
        if rect_x + w > ANCHO - 10:
            rect_x = max(10, ANCHO - w - 10)
        if rect_x < 10:
            rect_x = 10

        rect = pygame.Rect(rect_x, int(y), w, h)

        # Animación de fade-in / fade-out
        fade_in = 0.12
        fade_out = min(0.3, p['duration'] * 0.25)
        alpha_factor = 1.0
        if age < fade_in and fade_in > 0:
            alpha_factor = age / fade_in
        elif age > (p['duration'] - fade_out) and fade_out > 0:
            alpha_factor = max(0.0, (p['duration'] - age) / fade_out)
        alpha255 = int(220 * alpha_factor)

        # fondo semi-transparente con border-radius
        s = pygame.Surface((w, h), pygame.SRCALPHA)
        try:
            pygame.draw.rect(s, (*p['color'], alpha255), s.get_rect(), border_radius=8)
        except Exception:
            s.fill((*p['color'], alpha255))
        ventana.blit(s, rect.topleft)

        # dibujar icono si existe
        text_x = rect.x + padding_x
        if icon:
            icon_x = rect.x + padding_x
            icon_y = rect.y + padding_y + (h - padding_y*2 - icon_size)//2
            # círculo de fondo
            icon_surf = pygame.Surface((icon_size, icon_size), pygame.SRCALPHA)
            icon_surf.fill((0,0,0,0))
            pygame.draw.circle(icon_surf, (255,255,255, alpha255), (icon_size//2, icon_size//2), icon_size//2)
            # dibujar check si es 'check'
            if icon == 'check':
                # líneas del check
                cx = icon_size//2
                cy = icon_size//2
                pygame.draw.line(icon_surf, (30,160,60, alpha255), (4, icon_size//2), (icon_size//2, icon_size-5), 3)
                pygame.draw.line(icon_surf, (30,160,60, alpha255), (icon_size//2, icon_size-5), (icon_size-4, 6), 3)
            ventana.blit(icon_surf, (icon_x, icon_y))
            text_x += icon_space

        # texto (aplicar alpha a las superficies de texto)
        y_off = rect.y + padding_y
        for surf in line_surfs:
            surf.set_alpha(int(255 * alpha_factor))
            ventana.blit(surf, (text_x, y_off))
            y_off += surf.get_height() + 6
    # limpiar expirados (desde el final para no desordenar índices)
    for i in reversed(to_remove):
        floating_popups.pop(i)

def draw_right_panel():
    # Fondo azul ya dibujado
    padding = 20
    x0 = LEFT_WIDTH + padding
    y0 = padding

    # Título
    title = large_font.render(f'Nivel {level}', True, BLANCO)
    ventana.blit(title, (x0, y0))

    # Mostrar matriz (números)
    y_matrix = y0 + 50
    cell_h = 24
    for i in range(FILAS):
        # Convertir valores None a "#" y el resto a strings
        row_text = ' '.join('#' if tablero[i][j] is None else str(tablero[i][j]) for j in range(COLUMNAS))
        surf = font.render(row_text, True, BLANCO)
        ventana.blit(surf, (x0, y_matrix + i * (cell_h+2)))

    # Estadísticas
    y_stats = y_matrix + FILAS * (cell_h+2) + 20
    stats = [
        f'Puntos: {score}',
        f'Movimientos: {moves_count}',
        f'Tiempo restante: {max(0, int(LEVEL_TIME - (time.time() - level_start_time)))}s',
        f'Objetivo: {goal_for_level()}'
    ]
    for i, s in enumerate(stats):
        ventana.blit(font.render(s, True, BLANCO), (x0, y_stats + i * 26))

    # Botón HABILIDADES MAGICAS
    btn_rect = pygame.Rect(LEFT_WIDTH + 40, ALTO - 140, RIGHT_PANEL_WIDTH - 80, 40)
    pygame.draw.rect(ventana, GRIS, btn_rect, border_radius=6)
    txt = font.render('HABILIDADES MAGICAS', True, (0, 0, 0))
    ventana.blit(txt, (btn_rect.x + 12, btn_rect.y + 8))

    # Panel de estado de habilidades
    status_rect = pygame.Rect(LEFT_WIDTH + 20, y_stats + 120, RIGHT_PANEL_WIDTH - 40, 80)
    pygame.draw.rect(ventana, (0, 0, 0, 30), status_rect, border_radius=6)
    
    # Título del panel
    ventana.blit(font.render('Estado de Habilidades:', True, BLANCO), 
                (status_rect.x + 10, status_rect.y + 8))
    
    # Mostrar cooldowns y usos en el panel
    y_offset = status_rect.y + 32
    if quiz_cooldown_until > time.time():
        rem = int(quiz_cooldown_until - time.time())
        cd_txt = font.render(f'⏳ Penalización: {rem}s', True, BLANCO)
        ventana.blit(cd_txt, (status_rect.x + 10, y_offset))
        y_offset += 22
    
    if skill_success_cooldown_until > time.time():
        rem2 = int(skill_success_cooldown_until - time.time())
        cd_txt2 = font.render(f'⌛ Próxima habilidad: {rem2}s', True, BLANCO)
        ventana.blit(cd_txt2, (status_rect.x + 10, y_offset))
        y_offset += 22
    
    # Mostrar usos restantes por nivel
    try:
        usos_restantes = MAX_SKILL_USES_PER_LEVEL - skill_uses_this_level
    except NameError:
        usos_restantes = MAX_SKILL_USES_PER_LEVEL
    usos_txt = font.render(f'🎯 Usos: {skill_uses_this_level}/{MAX_SKILL_USES_PER_LEVEL}', True, BLANCO)
    ventana.blit(usos_txt, (status_rect.x + 10, y_offset))
        
    # Botón EXPORTAR DATOS
    export_btn = pygame.Rect(LEFT_WIDTH + 40, ALTO - 80, RIGHT_PANEL_WIDTH - 80, 40)
    pygame.draw.rect(ventana, (220, 220, 220), export_btn, border_radius=6)
    txt = font.render('EXPORTAR A EXCEL', True, (0, 0, 0))
    ventana.blit(txt, (export_btn.x + 12, export_btn.y + 8))

    # Botón ZONA DE ESTUDIO (learning en español)
    learn_btn = pygame.Rect(LEFT_WIDTH + 40, ALTO - 200, RIGHT_PANEL_WIDTH - 80, 40)
    pygame.draw.rect(ventana, (200, 230, 255), learn_btn, border_radius=6)
    txt = font.render('ZONA DE ESTUDIO', True, (0, 0, 0))
    ventana.blit(txt, (learn_btn.x + 12, learn_btn.y + 8))

def goal_for_level():
    # Objetivo que se duplica cada nivel (progresión exponencial 2^n)
    return int(goal_base * (2 ** (level - 1)))


def show_level_up(new_level):
    # Antes: modal bloqueante. Ahora crear popup flotante no bloqueante.
    lines = [f'¡Felicidades! Subiste al nivel {new_level}', f'Nuevo objetivo: {goal_base * (2 ** (new_level - 1))} pts']
    popup = {
        'lines': lines,
        'start': time.time(),
        'duration': 3.0,
        'x': ANCHO//2 - 180,
        'y': ALTO//2 - 40,
        'vy': -30.0,  # velocidad hacia arriba (pixeles por segundo)
        'color': (30, 120, 220)
    }
    floating_popups.append(popup)
    return

def obtener_celda(pos):
    x, y = pos
    if x < 0 or x >= LEFT_WIDTH:
        return None
    fila = (y - MARGEN_SUPERIOR) // TAMANO_CELDA
    columna = (x - MARGEN_IZQUIERDO) // TAMANO_CELDA
    if 0 <= fila < FILAS and 0 <= columna < COLUMNAS:
        return (fila, columna)
    return None

def son_adyacentes(p1, p2):
    (r1, c1), (r2, c2) = p1, p2
    return abs(r1 - r2) + abs(c1 - c2) == 1

def intercambiar(p1, p2):
    global moves_count
    r1, c1 = p1
    r2, c2 = p2
    
    # Guardar estado actual en historial
    current_matrix = [[tablero[i][j] for j in range(COLUMNAS)] for i in range(FILAS)]
    matrix_history.append(current_matrix)
    game_events.append(f"Intercambio ({r1},{c1}) con ({r2},{c2})")
    # intercambiar habilidades si existen (permitir mover la habilidad con el caramelo)
    has1 = p1 in skills_on_map
    has2 = p2 in skills_on_map
    if has1 or has2:
        v1 = skills_on_map.pop(p1) if has1 else None
        v2 = skills_on_map.pop(p2) if has2 else None
        if v1 is not None:
            skills_on_map[p2] = v1
        if v2 is not None:
            skills_on_map[p1] = v2
    # intercambiar colores del tablero
    tablero[r1][c1], tablero[r2][c2] = tablero[r2][c2], tablero[r1][c1]
    moves_count += 1

def find_matches():
    remove = set()
    # Horizontal
    for i in range(FILAS):
        run_color = tablero[i][0]
        run_start = 0
        for j in range(1, COLUMNAS + 1):
            color = tablero[i][j] if j < COLUMNAS else None
            if color == run_color:
                continue
            else:
                length = j - run_start
                if run_color is not None and length >= 3:
                    for k in range(run_start, j):
                        remove.add((i, k))
                if j < COLUMNAS:
                    run_color = tablero[i][j]
                    run_start = j

    # Vertical
    for j in range(COLUMNAS):
        run_color = tablero[0][j]
        run_start = 0
        for i in range(1, FILAS + 1):
            color = tablero[i][j] if i < FILAS else None
            if color == run_color:
                continue
            else:
                length = i - run_start
                if run_color is not None and length >= 3:
                    for k in range(run_start, i):
                        remove.add((k, j))
                if i < FILAS:
                    run_color = tablero[i][j]
                    run_start = i

    return remove

def remove_and_collapse(matches):
    global score
    if not matches:
        return 0
    # Reproducir sonido
    sound_manager.play_sound('explosion')

    removed = len(matches)
    score += removed * 10

    # Guardar para animación breve
    last_explosions.append((set(matches), time.time()))

    # Crear una copia del tablero para la animación
    tablero_anim = [[tablero[i][j] for j in range(COLUMNAS)] for i in range(FILAS)]
    # Eliminar marcando con None
    for (i, j) in matches:
        tablero[i][j] = None
        tablero_anim[i][j] = None

    # Calcular posiciones finales después del colapso
    tablero_final = [[None for _ in range(COLUMNAS)] for _ in range(FILAS)]
    for j in range(COLUMNAS):
        stack = [tablero[i][j] for i in range(FILAS) if tablero[i][j] is not None]
        # Rellenar arriba con nuevos
        while len(stack) < FILAS:
            stack.insert(0, random.randrange(NUM_COLORS))
        # Asignar a posiciones finales
        for i in range(FILAS):
            tablero_final[i][j] = stack[i]

    # Animación suave de caída
    start_time = time.time()
    duration = 0.5  # medio segundo de animación
    anim_clock = pygame.time.Clock()
    
    while time.time() - start_time < duration:
        progress = min(1.0, (time.time() - start_time) / duration)
        # Interpolar posiciones
        for j in range(COLUMNAS):
            col_stack = []
            # Encontrar caramelos existentes y sus destinos
            for i in range(FILAS):
                if tablero_anim[i][j] is not None:
                    # Buscar su posición final
                    found = False
                    for k in range(i, FILAS):
                        if tablero_final[k][j] == tablero_anim[i][j]:
                            found = True
                            # Interpolar posición
                            start_y = i
                            end_y = k
                            current_y = start_y + (end_y - start_y) * progress
                            col_stack.append((tablero_anim[i][j], current_y))
                            break
                    if not found and tablero_anim[i][j] is not None:
                        # El caramelo desaparece, moverlo hacia abajo
                        current_y = i + progress * 2
                        if current_y < FILAS:
                            col_stack.append((tablero_anim[i][j], current_y))
            
            # Dibujar nuevos caramelos cayendo desde arriba
            for i in range(FILAS):
                if tablero_final[i][j] not in [x[0] for x in col_stack]:
                    current_y = -1 + (i + 1) * progress
                    if current_y >= 0:
                        col_stack.append((tablero_final[i][j], current_y))
            
            # Actualizar tablero_anim para esta columna
            for i in range(FILAS):
                tablero_anim[i][j] = None
            for valor, y in col_stack:
                if 0 <= int(y) < FILAS:
                    tablero_anim[int(y)][j] = valor

        # Dibujar frame
        dibujar_interface()
        pygame.display.flip()
        anim_clock.tick(60)

    # Actualizar tablero final
    for i in range(FILAS):
        for j in range(COLUMNAS):
            tablero[i][j] = tablero_final[i][j]

    return removed

def spawn_skill_random(typ):
    """Genera una habilidad especial en una posición aleatoria del tablero.
    
    Args:
        typ (str): Tipo de habilidad (debe ser uno de Habilidad.get_all_types())
    
    Returns:
        bool: True si se pudo colocar la habilidad, False si no hay espacio disponible
    """
    if typ not in Habilidad.get_all_types():
        raise ValueError(f"Tipo de habilidad inválido: {typ}")
        
    # Obtener todas las celdas disponibles
    celdas_disponibles = [
        (r, c) for r in range(FILAS) for c in range(COLUMNAS)
        if (r, c) not in skills_on_map
    ]
    
    if not celdas_disponibles:
        return False
        
    # Elegir una celda aleatoria
    r, c = random.choice(celdas_disponibles)
    
    # Guardar el color original antes de reemplazar
    orig = tablero[r][c]
    skills_on_map[(r, c)] = (typ, orig)
    tablero[r][c] = None
    return True

def activate_skill_at(pos):
    global score
    if pos not in skills_on_map:
        return
    val = skills_on_map.pop(pos)
    if isinstance(val, tuple):
        typ, orig = val
    else:
        typ = val
        orig = None
    r, c = pos
    affected = set()
    if typ == 'bomb':
        for i in range(r-1, r+2):
            for j in range(c-1, c+2):
                if 0 <= i < FILAS and 0 <= j < COLUMNAS:
                    affected.add((i, j))
    elif typ == 'rainbow':
        for i in range(FILAS):
            affected.add((i, c))
        for j in range(COLUMNAS):
            affected.add((r, j))
    elif typ == 'star':
        # Usar el color original guardado cuando la habilidad fue colocada
        color = orig if orig is not None else tablero[r][c]
        for i in range(FILAS):
            for j in range(COLUMNAS):
                if tablero[i][j] == color:
                    affected.add((i, j))

    remove_and_collapse(affected)

def handle_quiz():
    global quiz_cooldown_until
    # Simple modal de preguntas: pregunta aleatoria entre la lista
    # Elegir pregunta sin repetir hasta agotar
    available = [i for i in range(len(questions)) if i not in asked_questions]
    if not available:
        asked_questions.clear()
        available = list(range(len(questions)))
    q_idx = random.choice(available)
    asked_questions.add(q_idx)
    pregunta, respuesta = questions[q_idx]
    
    # Ahora soportamos tres tipos de pregunta:
    # - TF: pregunta/TrueFalse (desde lista `questions`)
    # - MCQ: estructura con 'lines','options','correct_idx'
    # - MATRIX: generado dinámicamente en generate_matrix_mcq() (devuelve MCQ)
    # Elegir tipo con probabilidades (mezcla de TF y MCQ/matrix)
    q_struct = None
    pick = random.random()
    if pick < 0.5:
        # TF - usar pregunta existente
        available = [i for i in range(len(questions)) if i not in asked_questions]
        if not available:
            asked_questions.clear()
            available = list(range(len(questions)))
        q_idx = random.choice(available)
        asked_questions.add(q_idx)
        pregunta, respuesta = questions[q_idx]
        lines = []
        palabras = pregunta.split()
        cur = palabras[0] if palabras else ''
        for w in palabras[1:]:
            if len(cur + ' ' + w) <= 40:
                cur += ' ' + w
            else:
                lines.append(cur)
                cur = w
        if cur:
            lines.append(cur)
        # opciones Verdadero / Falso
        opts = ['Verdadero', 'Falso']
        correct_idx = 0 if respuesta else 1
        q_struct = {'lines': lines, 'options': opts, 'correct_idx': correct_idx}
    elif pick < 0.8:
        # Matrix-based MCQ
        q_struct = generate_matrix_mcq()
    else:
        # MCQ genérica (podríamos ampliar con preguntas creadas manualmente)
        # Ejemplo simple: definición
        q = '¿Qué representa la traza de una matriz?'
        opts = ['Suma de autovalores', 'Producto de autovalores', 'Número de filas']
        correct_idx = 0
        lines = [q]
        q_struct = {'lines': lines, 'options': opts, 'correct_idx': correct_idx}

    asking = True
    clock = pygame.time.Clock()
    global skill_success_cooldown_until, skill_uses_this_level
    while asking:
        # Calcular posición para la ventana de preguntas
        y_start = 350  # Posición fija debajo de la matriz
        
        # Modal más compacto y mejor posicionado
        modal = pygame.Rect(LEFT_WIDTH + 20, y_start, RIGHT_PANEL_WIDTH - 40, 160)
        # Opciones estarán dentro del modal, debajo del bloque de pregunta
        options_base_x = modal.x + 10
        options_base_y = modal.y + 60  # reducir espacio vertical
        w_btn = modal.width - 20
        h_btn = 36
        gap = 8

        for ev in pygame.event.get():
            if ev.type == pygame.QUIT:
                pygame.quit(); sys.exit()
            if ev.type == pygame.MOUSEBUTTONDOWN:
                mx, my = pygame.mouse.get_pos()
                # botones de opciones (dentro del modal)
                btns = []
                for k in range(len(q_struct['options'])):
                    rect = pygame.Rect(options_base_x, options_base_y + k*(h_btn+gap), w_btn, h_btn)
                    btns.append(rect)
                for idx, rect in enumerate(btns):
                    if rect.collidepoint(mx, my):
                        if idx == q_struct['correct_idx']:
                            # Correct
                            spawn_skill_random(random.choice(['bomb','rainbow','star']))
                            skill_uses_this_level += 1
                            skill_success_cooldown_until = time.time() + 10
                        else:
                            quiz_cooldown_until = time.time() + 20
                        asking = False

        # Dibujar modal sobre la pantalla (dentro del panel derecho)
        dibujar_interface()
        pygame.draw.rect(ventana, (245,245,245), modal, border_radius=8)
        pygame.draw.rect(ventana, (200,200,200), modal, 2, border_radius=8)
        # Renderizar pregunta
        y_offset = 10
        for linea in q_struct['lines']:
            qsurf = font.render(linea, True, (0,0,0))
            ventana.blit(qsurf, (modal.x + 10, modal.y + y_offset))
            y_offset += 26

        # Dibujar opciones dentro del modal (no destacar la correcta)
        for k, opt in enumerate(q_struct['options']):
            rect = pygame.Rect(options_base_x, options_base_y + k*(h_btn+gap), w_btn, h_btn)
            pygame.draw.rect(ventana, (220,220,220), rect, border_radius=6)
            ventana.blit(font.render(opt, True, (0,0,0)), (rect.x + 10, rect.y + 8))

        pygame.display.flip()
        clock.tick(30)

def sum_matrix():
    s = 0
    for i in range(FILAS):
        for j in range(COLUMNAS):
            # tablero puede contener None o -1 para habilidades; ignorar
            v = tablero[i][j]
            if isinstance(v, int) and v >= 0:
                s += v
    return s

def matrix_stats():
    vals = []
    for i in range(FILAS):
        for j in range(COLUMNAS):
            v = tablero[i][j]
            if isinstance(v, int) and v >= 0:
                vals.append(v)
    if not vals:
        return {'sum':0, 'min':0, 'max':0, 'mean':0}
    return {'sum': sum(vals), 'min': min(vals), 'max': max(vals), 'mean': sum(vals)/len(vals)}

def show_end_screen(victory=False):
    # Mostrar ventana final compacta con estadísticas detalladas
    total_time = int(time.time() - game_start_time)
    stats = matrix_stats()
    clock = pygame.time.Clock()
    while True:
        for ev in pygame.event.get():
            if ev.type == pygame.QUIT:
                pygame.quit(); sys.exit()
            if ev.type == pygame.KEYDOWN or ev.type == pygame.MOUSEBUTTONDOWN:
                return

        # Fondo tenue
        ventana.fill((30,30,30))

        # Dibujar ventana central pequeña
        w = 520
        h = 360
        rx = (ANCHO - w)//2
        ry = (ALTO - h)//2
        rect = pygame.Rect(rx, ry, w, h)
        pygame.draw.rect(ventana, (235, 245, 255), rect, border_radius=12)

        # Título
        if victory:
            title = large_font.render('¡Felicidades! Juego completado', True, (10,90,30))
        else:
            title = large_font.render('Game Over', True, (150,10,10))
        ventana.blit(title, (rect.x + 20, rect.y + 16))

        # Estadísticas
        lines = [
            f'Nivel alcanzado: {level}',
            f'Puntos totales: {score}',
            f'Movimientos totales: {moves_count}',
            f'Tiempo total (s): {total_time}',
            f'Suma matriz: {stats["sum"]}',
            f'Media matriz: {stats["mean"]:.2f}',
            f'Mín matriz: {stats["min"]}  Máx matriz: {stats["max"]}'
        ]
        for i, l in enumerate(lines):
            ventana.blit(font.render(l, True, (20,20,20)), (rect.x + 24, rect.y + 80 + i*32))

        hint = font.render('Haz click o presiona cualquier tecla para cerrar', True, (80,80,80))
        ventana.blit(hint, (rect.x + 24, rect.y + h - 40))

        pygame.display.flip()
        clock.tick(30)


# Inicialización de variables del juego
score = 0
level = 1
moves_count = 0
seleccionado = None
tablero = [[random.randrange(NUM_COLORS) for _ in range(COLUMNAS)] for _ in range(FILAS)]
matrix_history = []  # historial de matrices para exportar
game_events = []     # eventos del juego para exportar

# Bucle principal
clock = pygame.time.Clock()
running = True
while running:
    for evento in pygame.event.get():
        if evento.type == pygame.QUIT:
            running = False
        elif evento.type == pygame.MOUSEBUTTONDOWN:
            pos = obtener_celda(pygame.mouse.get_pos())
            mx, my = pygame.mouse.get_pos()
            # Verificar botón preguntas
            btn_rect = pygame.Rect(LEFT_WIDTH + 40, ALTO - 140, RIGHT_PANEL_WIDTH - 80, 40)
            export_btn = pygame.Rect(LEFT_WIDTH + 40, ALTO - 80, RIGHT_PANEL_WIDTH - 80, 40)
            
            if btn_rect.collidepoint(mx, my):
                now = time.time()
                # Priorizar penalización por respuesta incorrecta
                if now < quiz_cooldown_until:
                    rem = int(quiz_cooldown_until - now)
                    popup = {
                        'lines': [f'No puedes usar ahora las habilidades.', f'Penalización activa: {rem}s'],
                        'start': time.time(),
                        'duration': 2.5,
                        'x': ANCHO//2 - 180,
                        'y': ALTO//2 - 40,
                        'vy': -20.0,
                        'color': (180, 60, 60)
                    }
                    floating_popups.append(popup)
                elif skill_uses_this_level >= MAX_SKILL_USES_PER_LEVEL:
                    popup = {
                        'lines': [f'Has usado el máximo de habilidades en este nivel ({MAX_SKILL_USES_PER_LEVEL}).'],
                        'start': time.time(),
                        'duration': 2.5,
                        'x': ANCHO//2 - 180,
                        'y': ALTO//2 - 40,
                        'vy': -20.0,
                        'color': (200, 140, 40)
                    }
                    floating_popups.append(popup)
                elif now < skill_success_cooldown_until:
                    rem = int(skill_success_cooldown_until - now)
                    popup = {
                        'lines': [f'Habilidad en cooldown. Espera {rem}s.'],
                        'start': time.time(),
                        'duration': 2.5,
                        'x': ANCHO//2 - 180,
                        'y': ALTO//2 - 40,
                        'vy': -20.0,
                        'color': (100, 120, 200)
                    }
                    floating_popups.append(popup)
                else:
                    handle_quiz()
            # LEARNING button area (same vertical spacing as render)
            learn_btn = pygame.Rect(LEFT_WIDTH + 40, ALTO - 200, RIGHT_PANEL_WIDTH - 80, 40)
            if learn_btn.collidepoint(mx, my):
                show_learn_screen()
            elif export_btn.collidepoint(mx, my):
                if EXCEL_AVAILABLE:
                    filename = export_to_excel()
                    # Mostrar mensaje de éxito como popup flotante
                    popup = {
                        'lines': [f'Datos exportados a:', filename],
                        'start': time.time(),
                        'duration': 3.0,
                        'x': ANCHO//2 - 180,
                        'y': ALTO//2 - 40,
                        'vy': -30.0,
                        'color': (30, 180, 30),
                        'icon': 'check'
                    }
                    floating_popups.append(popup)
            else:
                if pos:
                    if not seleccionado:
                        seleccionado = pos
                    else:
                        if son_adyacentes(seleccionado, pos):
                            intercambiar(seleccionado, pos)
                            # Activar habilidad si se movió sobre una
                            if pos in skills_on_map:
                                activate_skill_at(pos)
                            # Buscar matches y resolver cascadas
                            total_removed = 0
                            while True:
                                matches = find_matches()
                                if not matches:
                                    break
                                removed = remove_and_collapse(matches)
                                total_removed += removed
                            seleccionado = None
                        else:
                            seleccionado = pos
        elif evento.type == pygame.MOUSEMOTION:
            mx, my = pygame.mouse.get_pos()
            hover_cell = obtener_celda((mx, my))

    # Comprobar si se alcanzó objetivo para subir de nivel
    if score >= goal_for_level():
        if level >= MAX_LEVEL:
            # completó el último nivel: victoria
            show_end_screen(victory=True)
            running = False
            break
        else:
            # informar subida de nivel
            show_level_up(level + 1)
            level += 1
            # Resetar contador de usos y cooldowns al subir de nivel
            try:
                skill_uses_this_level = 0
                skill_success_cooldown_until = 0
                quiz_cooldown_until = 0
            except NameError:
                pass
            level_start_time = time.time()

    # Comprobar tiempo de nivel
    elapsed = time.time() - level_start_time
    if elapsed >= LEVEL_TIME:
        # tiempo finalizado
        if score >= goal_for_level():
            if level >= MAX_LEVEL:
                # Ganó el juego
                # mostrar pantalla de felicitaciones simple y salir
                ventana.fill(BLANCO)
                msg = large_font.render('¡Felicidades! Completaste el nivel final.', True, (0,0,0))
                ventana.blit(msg, (50, ALTO//2 - 20))
                pygame.display.flip()
                pygame.time.delay(3000)
                running = False
            else:
                level += 1
                # Resetar contador de usos y cooldowns al subir de nivel
                try:
                    skill_uses_this_level = 0
                    skill_success_cooldown_until = 0
                    quiz_cooldown_until = 0
                except NameError:
                    pass
                level_start_time = time.time()
                # aumentar objetivo y continuar
        else:
            # Game over
            show_end_screen(victory=False)
            running = False

    dibujar_interface()
    pygame.display.flip()
    clock.tick(30)

pygame.quit()
sys.exit()